import os
import sys
import tempfile
from pathlib import Path
from typing import List, Optional, Tuple, Dict
import re

import streamlit as st
import pdfplumber
import fitz  # PyMuPDF
from PIL import Image
import pytesseract

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, DataBarRule
from openpyxl.styles.colors import Color
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.cell.cell import MergedCell

# =========================
# CONFIG TESSERACT
# =========================
def app_base_path() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR = app_base_path()
TESSERACT_EXE = os.path.join(BASE_DIR, "Tesseract-OCR", "tesseract.exe")
TESSDATA_DIR = os.path.join(BASE_DIR, "Tesseract-OCR", "tessdata")

if os.path.exists(TESSERACT_EXE):
    pytesseract.pytesseract.tesseract_cmd = TESSERACT_EXE
if os.path.exists(TESSDATA_DIR):
    os.environ["TESSDATA_PREFIX"] = TESSDATA_DIR


# =========================
# HELPERS
# =========================
def parse_bs_number(s: str) -> Optional[int]:
    if s is None:
        return None
    s = str(s).strip()
    if not s or s == "-":
        return None

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1]

    s = s.replace(" ", "").replace(".", "").replace(",", "")
    if not s.isdigit():
        return None

    val = int(s)
    return -val if negative else val


def parse_pages_arg(s: str) -> List[int]:
    s = s.strip()
    parts = [p.strip() for p in s.split(",") if p.strip()]
    pages: List[int] = []

    for p in parts:
        if "-" in p:
            a, b = p.split("-", 1)
            a, b = int(a), int(b)
            if a > b:
                a, b = b, a
            pages.extend(list(range(a, b + 1)))
        else:
            pages.append(int(p))

    seen = set()
    out = []
    for x in pages:
        if x not in seen:
            seen.add(x)
            out.append(x)
    return out


def is_num_token(t: str) -> bool:
    t = t.strip()
    if not t:
        return False
    return bool(re.fullmatch(r"-?\(?\d[\d\.\,]*\)?", t))


def rows_from_text_lines(lines: List[str], mode: str = "default") -> List[Tuple[str, Optional[int], Optional[int]]]:
    rows: List[Tuple[str, Optional[int], Optional[int]]] = []

    TITLE_PHRASES = [
        "POR LOS EJERCICIOS",
        "EJERCICIOS TERMINADOS",
        "TERMINADOS EL",
        "AL 31 DE DICIEMBRE",
        "DEL 31 DE DICIEMBRE",
        "AL31 DE DICIEMBRE",
        "LAS NOTAS",
        "PARTE INTEGRANTE",
        "ESTADOS FINANCIEROS",
        "NOTAS ADJ",
    ]

    num_re = re.compile(r"\(?\d[\d\.\,]*\)?")

    col23_pos = None
    col22_pos = None
    if mode == "flujo":
        p23 = []
        p22 = []
        for ln in lines:
            s = (ln or "").strip()
            if not s:
                continue
            up = s.upper()

            if any(p in up for p in TITLE_PHRASES):
                continue
            if "NOTAS" in up and ("2023" in up or "2022" in up):
                continue
            if "CUENTA" in up and ("2023" in up or "2022" in up):
                continue

            ms = list(num_re.finditer(s))
            if len(ms) >= 2:
                p23.append(ms[-2].start())
                p22.append(ms[-1].start())

        if p23 and p22:
            p23.sort()
            p22.sort()
            col23_pos = p23[len(p23) // 2]
            col22_pos = p22[len(p22) // 2]

    i = 0
    while i < len(lines):
        line = (lines[i] or "").strip()
        if not line:
            i += 1
            continue

        up = line.upper()

        if any(p in up for p in TITLE_PHRASES):
            i += 1
            continue
        if "NOTAS" in up and ("2023" in up or "2022" in up):
            i += 1
            continue
        if "CUENTA" in up and ("2023" in up or "2022" in up):
            i += 1
            continue

        tokens = line.split()

        if len(tokens) <= 2 and not any(is_num_token(t) for t in tokens):
            i += 1
            continue
        if len(tokens) == 1 and len(tokens[0]) <= 5:
            i += 1
            continue

        num_positions = [idx for idx, tok in enumerate(tokens) if is_num_token(tok)]
        if len(num_positions) == 0:
            i += 1
            continue

        num_tokens = [tokens[j].strip("()") for j in num_positions]
        only_years = (len(num_tokens) == 2 and set(num_tokens) <= {"2023", "2022"})
        if only_years:
            i += 1
            continue

        matches = list(num_re.finditer(line))

        v23 = None
        v22 = None

        if len(num_positions) >= 2:
            i23, i22 = num_positions[-2], num_positions[-1]
            v23 = parse_bs_number(tokens[i23])
            v22 = parse_bs_number(tokens[i22])
            cuenta = " ".join(tokens[:i23]).strip()
        else:
            i1 = num_positions[-1]
            one_val = parse_bs_number(tokens[i1])
            cuenta = " ".join(tokens[:i1]).strip()

            if mode == "flujo" and col23_pos is not None and col22_pos is not None and len(matches) == 1:
                pos = matches[0].start()
                if abs(pos - col22_pos) < abs(pos - col23_pos):
                    v23 = None
                    v22 = one_val
                else:
                    v23 = one_val
                    v22 = None
            else:
                v23 = one_val
                v22 = None

            if mode == "flujo" and i + 1 < len(lines):
                nxt = (lines[i + 1] or "").strip()
                if nxt:
                    has_letters = bool(re.search(r"[A-Za-zÁÉÍÓÚÑáéíóúñ]", nxt))
                    nxt_matches = list(num_re.finditer(nxt))
                    if (not has_letters) and len(nxt_matches) == 1:
                        other_val = parse_bs_number(nxt_matches[0].group())
                        other_pos = nxt_matches[0].start()
                        if col23_pos is not None and col22_pos is not None:
                            if abs(other_pos - col22_pos) < abs(other_pos - col23_pos):
                                v22 = other_val if v22 is None else v22
                            else:
                                v23 = other_val if v23 is None else v23
                        else:
                            v22 = other_val if v22 is None else v22
                        i += 1

        cuenta = cuenta.replace("“", "").replace("”", "").replace('"', "").replace("'", "").strip()
        cuenta = re.sub(r"\s*(?:—|-)?\s*(?:\d+\.[A-Za-z0-9]\s*(?:y|&)\s*\d+\.[A-Za-z0-9])\s*$", "", cuenta, flags=re.IGNORECASE)
        cuenta = re.sub(r"\s*(?:—|-)?\s*\d+\.[A-Za-z0-9]\s*$", "", cuenta, flags=re.IGNORECASE)
        cuenta = re.sub(r"\s*[A-Za-z]{1,3}\s+\d+\.[A-Za-z0-9]\s*$", "", cuenta, flags=re.IGNORECASE)
        cuenta = re.sub(r"\s+(?:y|&)\s*$", "", cuenta, flags=re.IGNORECASE).strip()
        cuenta = cuenta.rstrip("—-– ").strip()
        cuenta = re.sub(r"\s*(?:—|-)?\s*\d{1,2}[A-Za-z]\s*$", "", cuenta)
        cuenta = re.sub(r"\s*(?:—|-)?\s*[A-Za-z]\d{1,2}\s*$", "", cuenta)
        cuenta = cuenta.replace("NOTAS", "").replace("NOTA", "").replace("NOTE", "").replace("NOTES", "").strip()

        cuenta_up = cuenta.upper().strip()
        if "CAUCRUZ" in cuenta_up or "SH PECKER" in cuenta_up:
            i += 1
            continue

        if not cuenta or len(cuenta) <= 2:
            i += 1
            continue

        cuenta_norm = re.sub(r"\s+", "", cuenta)
        if re.fullmatch(r"[^\w]*[A-Za-z]{1,2}[^\w]*", cuenta_norm):
            i += 1
            continue

        if mode == "default":
            if len(cuenta) <= 3 and (v23 is not None and abs(v23) < 100000) and v22 is None:
                i += 1
                continue

        rows.append((cuenta, v23, v22))
        i += 1

    seen = set()
    cleaned = []
    for r in rows:
        if r in seen:
            continue
        seen.add(r)
        cleaned.append(r)

    return cleaned


def extract_page_lines_text(pdf_path: Path, page_1based: int) -> List[str]:
    try:
        with pdfplumber.open(str(pdf_path)) as pdf:
            idx = page_1based - 1
            if idx < 0 or idx >= len(pdf.pages):
                return []
            page = pdf.pages[idx]
            txt = page.extract_text() or ""
            txt = txt.strip()
            if txt:
                return txt.splitlines()
    except Exception:
        pass

    try:
        doc = fitz.open(str(pdf_path))
        idx = page_1based - 1
        if idx < 0 or idx >= doc.page_count:
            doc.close()
            return []

        page = doc.load_page(idx)
        mat = fitz.Matrix(2.5, 2.5)
        pix = page.get_pixmap(matrix=mat, alpha=False)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        ocr_text = pytesseract.image_to_string(img, lang="spa+eng", config="--psm 6")
        doc.close()

        ocr_text = (ocr_text or "").strip()
        if not ocr_text:
            return []
        return ocr_text.splitlines()
    except Exception:
        return []


def extract_section(pdf_path: Path, pages_1based: List[int], mode: str = "default") -> List[Tuple[str, Optional[int], Optional[int]]]:
    all_lines: List[str] = []
    for p in pages_1based:
        all_lines.extend(extract_page_lines_text(pdf_path, p))

    cleaned_lines = []
    for ln in all_lines:
        ln = (ln or "").strip()
        if len(ln) < 2:
            continue
        cleaned_lines.append(ln)

    return rows_from_text_lines(cleaned_lines, mode=mode)


# =========================
# EXCEL HELPERS
# =========================
def apply_table_borders(ws):
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).border = border


def upsert_sheet(wb, sheet_name: str, rows: List[Tuple[str, Optional[int], Optional[int]]]):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(sheet_name)

    ws.append(["Cuenta", "2023 (Bs)", "2022 (Bs)"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cuenta, v23, v22 in rows:
        ws.append([cuenta, v23, v22])

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15

    num_fmt = '#,##0;[Red]-#,##0'
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=2).number_format = num_fmt
        ws.cell(row=r, column=3).number_format = num_fmt

    def should_bold(c: str) -> bool:
        c_up = (c or "").strip().upper().lstrip("_").rstrip("=").strip()
        if sheet_name == "Balance General":
            return c_up.startswith("TOTAL")
        if sheet_name == "Estado de Resultado":
            return c_up.startswith("UTILIDAD") or c_up.startswith("RESULTADO")
        if sheet_name == "Flujo de Caja":
            return c_up.startswith("TOTAL") or "DISPONIBILIDAD AL CIERRE" in c_up
        return False

    bold_font = Font(bold=True)
    for r in range(2, ws.max_row + 1):
        cuenta = ws.cell(row=r, column=1).value
        if should_bold(cuenta):
            ws.cell(row=r, column=1).font = bold_font
            ws.cell(row=r, column=2).font = bold_font
            ws.cell(row=r, column=3).font = bold_font

    apply_table_borders(ws)


def _normalize_key(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().upper())


def sheet_to_map(ws) -> Dict[str, Tuple[Optional[int], Optional[int]]]:
    m: Dict[str, Tuple[Optional[int], Optional[int]]] = {}
    for r in range(2, ws.max_row + 1):
        k = _normalize_key(ws.cell(r, 1).value)
        if not k:
            continue
        v23 = ws.cell(r, 2).value
        v22 = ws.cell(r, 3).value
        m[k] = (v23, v22)
    return m


def find_value(m: Dict[str, Tuple[Optional[int], Optional[int]]], keywords: List[str], exclude_contains: Optional[List[str]] = None) -> Tuple[Optional[int], Optional[int]]:
    exclude_contains = exclude_contains or []
    ex = [_normalize_key(x) for x in exclude_contains]
    keys = list(m.keys())

    def ok(k: str) -> bool:
        ku = _normalize_key(k)
        return not any(e in ku for e in ex)

    for kw in keywords:
        kwu = _normalize_key(kw)
        for k in keys:
            if ok(k) and _normalize_key(k) == kwu:
                return m[k]
    for kw in keywords:
        kwu = _normalize_key(kw)
        for k in keys:
            if ok(k) and _normalize_key(k).startswith(kwu):
                return m[k]
    for kw in keywords:
        kwu = _normalize_key(kw)
        for k in keys:
            if ok(k) and kwu in _normalize_key(k):
                return m[k]
    return (None, None)


def safe_div(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None or b == 0:
        return None
    return a / b


def create_ratios_sheet(wb):
    if "Balance General" not in wb.sheetnames or "Estado de Resultado" not in wb.sheetnames:
        return

    bg = wb["Balance General"]
    er = wb["Estado de Resultado"]
    bg_map = sheet_to_map(bg)
    er_map = sheet_to_map(er)

    act_corr_23, act_corr_22 = find_value(bg_map, ["ACTIVO CORRIENTE", "TOTAL ACTIVO CORRIENTE"])
    inv_23, inv_22 = find_value(bg_map, ["INVENTARIOS", "INVENTARIO"])
    disp_23, disp_22 = find_value(bg_map, ["DISPONIBILIDADES", "EFECTIVO", "CAJA"])
    pas_corr_23, pas_corr_22 = find_value(bg_map, ["PASIVO CORRIENTE", "TOTAL PASIVO CORRIENTE"])
    pas_total_23, pas_total_22 = find_value(bg_map, ["TOTAL DEL PASIVO", "TOTAL PASIVO"], exclude_contains=["TOTAL PASIVO Y PATRIMONIO"])
    act_total_23, act_total_22 = find_value(bg_map, ["TOTAL ACTIVO"])
    pat_23, pat_22 = find_value(bg_map, ["TOTAL PATRIMONIO NETO", "TOTAL DEL PATRIMONIO NETO"], exclude_contains=["TOTAL PASIVO Y PATRIMONIO"])
    cxc_23, cxc_22 = find_value(bg_map, ["CUENTAS POR COBRAR"])
    cxp_23, cxp_22 = find_value(bg_map, ["PROVEEDORES", "CUENTAS POR PAGAR"])

    ventas_23, ventas_22 = find_value(er_map, ["INGRESOS POR VENTAS", "VENTAS"])
    costo_23, costo_22 = find_value(er_map, ["COSTO DE VENTAS", "COSTO"])
    util_op_23, util_op_22 = find_value(er_map, ["UTILIDAD OPERATIVA"])
    util_neta_23, util_neta_22 = find_value(er_map, ["UTILIDAD DEL EJERCICIO", "UTILIDAD NETA", "RESULTADO DEL EJERCICIO"])
    intereses_23, intereses_22 = find_value(er_map, ["INTERESES", "GASTOS FINANCIEROS"])

    ventas23 = abs(ventas_23) if ventas_23 is not None else None
    ventas22 = abs(ventas_22) if ventas_22 is not None else None
    costo23 = abs(costo_23) if costo_23 is not None else None
    costo22 = abs(costo_22) if costo_22 is not None else None
    act23 = abs(act_total_23) if act_total_23 is not None else None
    act22 = abs(act_total_22) if act_total_22 is not None else None
    pas23 = abs(pas_total_23) if pas_total_23 is not None else None
    pas22 = abs(pas_total_22) if pas_total_22 is not None else None
    pat23 = abs(pat_23) if pat_23 is not None else None
    pat22 = abs(pat_22) if pat_22 is not None else None
    un23 = abs(util_neta_23) if util_neta_23 is not None else None
    un22 = abs(util_neta_22) if util_neta_22 is not None else None
    util_bruta23 = None if (ventas23 is None or costo23 is None) else (ventas23 - costo23)
    util_bruta22 = None if (ventas22 is None or costo22 is None) else (ventas22 - costo22)

    ratios = [
        ("Liquidez Corriente", safe_div(act_corr_23, pas_corr_23), safe_div(act_corr_22, pas_corr_22)),
        ("Liquidez Ácida", safe_div(None if act_corr_23 is None or inv_23 is None else (act_corr_23 - inv_23), pas_corr_23),
         safe_div(None if act_corr_22 is None or inv_22 is None else (act_corr_22 - inv_22), pas_corr_22)),
        ("Liquidez Inmediata", safe_div(disp_23, pas_corr_23), safe_div(disp_22, pas_corr_22)),
        ("Índice de Deuda", safe_div(pas23, act23), safe_div(pas22, act22)),
        ("Índice de Calidad de Deuda", safe_div(pas_corr_23, pas_total_23), safe_div(pas_corr_22, pas_total_22)),
        ("Multiplicador del Apalancamiento", safe_div(act23, pat23), safe_div(act22, pat22)),
        ("Cobertura de Intereses", safe_div(util_op_23, intereses_23), safe_div(util_op_22, intereses_22)),
        ("Rotación de Activos", safe_div(ventas23, act23), safe_div(ventas22, act22)),
        ("Período Promedio de Cobro", None if safe_div(ventas_23, cxc_23) is None else abs(360.0 / safe_div(ventas_23, cxc_23)),
         None if safe_div(ventas_22, cxc_22) is None else abs(360.0 / safe_div(ventas_22, cxc_22))),
        ("Período Promedio de Inventario", None if safe_div(costo_23, inv_23) is None else abs(360.0 / safe_div(costo_23, inv_23)),
         None if safe_div(costo_22, inv_22) is None else abs(360.0 / safe_div(costo_22, inv_22))),
        ("Período Promedio de Pago", None if safe_div(costo_23, cxp_23) is None else abs(360.0 / safe_div(costo_23, cxp_23)),
         None if safe_div(costo_22, cxp_22) is None else abs(360.0 / safe_div(costo_22, cxp_22))),
        ("Ciclo de Conversión de Efectivo",
         None if any(v is None for v in [safe_div(costo_23, inv_23), safe_div(ventas_23, cxc_23), safe_div(costo_23, cxp_23)]) else abs(abs(360.0 / safe_div(costo_23, inv_23)) + abs(360.0 / safe_div(ventas_23, cxc_23)) - abs(360.0 / safe_div(costo_23, cxp_23))),
         None if any(v is None for v in [safe_div(costo_22, inv_22), safe_div(ventas_22, cxc_22), safe_div(costo_22, cxp_22)]) else abs(abs(360.0 / safe_div(costo_22, inv_22)) + abs(360.0 / safe_div(ventas_22, cxc_22)) - abs(360.0 / safe_div(costo_22, cxp_22)))),
        ("Margen Bruto", safe_div(util_bruta23, ventas23), safe_div(util_bruta22, ventas22)),
        ("Margen Operativo", safe_div(util_op_23, ventas_23), safe_div(util_op_22, ventas_22)),
        ("Margen Neto", safe_div(util_neta_23, ventas_23), safe_div(util_neta_22, ventas_22)),
        ("ROA (Rentabilidad sobre Activos)", safe_div(un23, act23), safe_div(un22, act22)),
        ("ROE (Rentabilidad sobre Patrimonio)", safe_div(util_neta_23, pat_23), safe_div(util_neta_22, pat_22)),
    ]

    name = "Ratios"
    if name in wb.sheetnames:
        ws = wb[name]
        ws.delete_rows(1, ws.max_row)
    else:
        ws = wb.create_sheet(name)

    def ratio_category(rname: str) -> str:
        r = (rname or "").upper()
        if "LIQUIDEZ" in r:
            return "LIQUIDEZ"
        if "DEUDA" in r or "COBERTURA" in r or "APALANCAMIENTO" in r:
            return "DEUDA"
        if "ROTACIÓN" in r or "PERÍODO" in r or "CICLO" in r:
            return "ROTACIÓN"
        if "MARGEN" in r or "ROA" in r or "ROE" in r or "RENTABILIDAD" in r:
            return "MÁRGENES"
        if "DUPONT" in r:
            return "ANÁLISIS DUPONT"
        return "OTROS"

    ws.append(["Categoría", "Ratio", "2023", "2022"])
    for c in ws[1]:
        c.font = Font(bold=True)

    for rname, r23, r22 in ratios:
        ws.append([ratio_category(rname), rname, r23, r22])

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 3).number_format = "0.00"
        ws.cell(r, 4).number_format = "0.00"

    apply_table_borders(ws)


# =========================
# DASHBOARD (simplificado, reutiliza tu versión estable)
# =========================
def create_dashboard_sheet(wb):
    # Si querés, acá después se puede pegar tu versión completa del dashboard
    pass


# =========================
# BUILD
# =========================
def build_excel(pdf_path: Path, balance_pages: List[int], er_pages: List[int], flujo_pages: List[int]) -> Path:
    data = {
        "Balance General": extract_section(pdf_path, balance_pages, mode="default"),
        "Estado de Resultado": extract_section(pdf_path, er_pages, mode="default"),
        "Flujo de Caja": extract_section(pdf_path, flujo_pages, mode="flujo"),
    }

    out_xlsx = pdf_path.with_suffix(".xlsx")

    if out_xlsx.exists():
        wb = load_workbook(out_xlsx)
    else:
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)

    upsert_sheet(wb, "Balance General", data["Balance General"])
    upsert_sheet(wb, "Estado de Resultado", data["Estado de Resultado"])
    upsert_sheet(wb, "Flujo de Caja", data["Flujo de Caja"])
    create_ratios_sheet(wb)
    # create_dashboard_sheet(wb)  # descomentá cuando pegues tu versión final estable
    wb.save(out_xlsx)
    return out_xlsx


# =========================
# STREAMLIT UI
# =========================
st.set_page_config(page_title="Extractor de Estados Financieros", layout="wide")
st.title("Extractor de Estados Financieros (PDF/OCR → Excel)")
st.caption("Sube un PDF, define páginas de Balance, Estado de Resultado y Flujo, y descarga el Excel.")

with st.sidebar:
    st.subheader("Configuración OCR")
    st.write(f"Tesseract detectado: {'Sí' if os.path.exists(TESSERACT_EXE) else 'No'}")
    st.write(f"Tessdata detectado: {'Sí' if os.path.exists(TESSDATA_DIR) else 'No'}")

uploaded_file = st.file_uploader("Sube tu PDF", type=["pdf"])

col1, col2, col3 = st.columns(3)
with col1:
    balance_pages = st.text_input("Páginas Balance", value="7")
with col2:
    er_pages = st.text_input("Páginas Estado de Resultado", value="8")
with col3:
    flujo_pages = st.text_input("Páginas Flujo de Caja", value="10")

run = st.button("Generar Excel", type="primary")

if run:
    if uploaded_file is None:
        st.error("Primero sube un PDF.")
    else:
        try:
            bal = parse_pages_arg(balance_pages)
            er = parse_pages_arg(er_pages)
            flujo = parse_pages_arg(flujo_pages)

            with tempfile.TemporaryDirectory() as tmpdir:
                pdf_path = Path(tmpdir) / uploaded_file.name
                pdf_path.write_bytes(uploaded_file.getbuffer())

                with st.spinner("Procesando PDF y generando Excel..."):
                    out_xlsx = build_excel(pdf_path, bal, er, flujo)
                    data = out_xlsx.read_bytes()

                st.success("Excel generado correctamente.")
                st.download_button(
                    label="Descargar Excel",
                    data=data,
                    file_name=out_xlsx.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        except Exception as e:
            st.exception(e)
